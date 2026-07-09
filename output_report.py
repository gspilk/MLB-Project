"""
output_report.py
Generates the Excel report from all analysis.

Sheets:
  1. Team Summary      — diagnosis + standings + outlook
  2. Player Grades     — every player graded
  3. Statcast          — xwOBA + luck analysis
  4. Stats to Improve  — ranked weaknesses + fixes
  5. Roster Moves      — immediate + deadline
  6. Schedule          — next 7 + remaining games

Usage:
    from output_report import generate_report
    generate_report(data, analysis, grades, recs)
"""

import os
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter

OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR,
    f"mariners_report_{date.today().strftime('%Y%m%d')}.xlsx")

# ── colors ────────────────────────────────────────────────────────────────────
NAVY     = "1B2A4A"
TEAL     = "005C5C"
GOLD     = "C4A535"
WHITE    = "FFFFFF"
LGRAY    = "F2F2F2"
DGRAY    = "CCCCCC"
GREEN    = "D4EDDA"
YELLOW   = "FFF3CD"
RED      = "F8D7DA"
ORANGE   = "FFE5CC"

GRADE_COLORS = {
    "Elite":                     GREEN,
    "Above Average":             "E8F5E9",
    "Average":                   LGRAY,
    "Below Average":             YELLOW,
    "Below Average — developing":YELLOW,
    "Small sample":              LGRAY,
    "DFA":                       RED,
    "Unknown":                   LGRAY,
    "Limited — injured/inactive":LGRAY,
}


# ── style helpers ─────────────────────────────────────────────────────────────
def _header(ws, row, col, value, bg=NAVY, fg=WHITE,
            bold=True, size=11, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center",
                               vertical="center",
                               wrap_text=wrap)
    return cell


def _cell(ws, row, col, value, bold=False, color=None,
          align="left", fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if color:
        cell.fill = PatternFill("solid", start_color=color)
    if fmt:
        cell.number_format = fmt
    return cell


def _title(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = Font(bold=True, size=14, color=WHITE, name="Arial")
    cell.fill      = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def _subtitle(ws, row, text, span=8, bg=TEAL):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = Font(bold=True, size=11, color=WHITE, name="Arial")
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               indent=1)
    ws.row_dimensions[row].height = 20


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _grade_color(grade: str) -> str:
    base = grade.split("—")[0].strip()
    return GRADE_COLORS.get(grade, GRADE_COLORS.get(base, LGRAY))


# ── sheet 1: team summary ─────────────────────────────────────────────────────
def _sheet_summary(wb: Workbook, d: dict, ou: dict):
    ws = wb.create_sheet("Team Summary")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [22, 18, 18, 18, 18, 18, 18, 18])

    r = 1
    _title(ws, r, f"SEATTLE MARINERS — MIDSEASON REPORT  |  {date.today()}", 8)
    r += 1

    # overall
    _subtitle(ws, r, "OVERALL GRADE", 8)
    r += 1
    labels = ["Grade", "Verdict", "Record", "Div Rank",
              "MLB Rank", "Proj Wins", "Floor", "Ceiling"]
    for i, l in enumerate(labels, 1):
        _header(ws, r, i, l, bg=TEAL, size=10)
    r += 1
    vals = [
        d.get("grade"), d.get("verdict"), d.get("record"),
        f"{d.get('div_rank')}/5", f"{d.get('mlb_rank')}/30",
        d.get("projected_wins"), d.get("floor"), d.get("ceiling")
    ]
    for i, v in enumerate(vals, 1):
        _cell(ws, r, i, v, bold=True, align="center",
              color=GREEN if d.get("grade","").startswith("A") else LGRAY)
    r += 2

    # standings
    _subtitle(ws, r, "STANDINGS & LUCK", 8)
    r += 1
    stat_rows = [
        ("Record",        d.get("record")),
        ("Win %",         d.get("win_pct")),
        ("Luck",          d.get("luck")),
        ("Pythagorean W-L",d.get("pythag")),
        ("1-Run Record",  d.get("one_run")),
        ("vs LHP",        d.get("vlhp")),
        ("Last 10",       d.get("last10")),
        ("Last 30",       d.get("last30")),
    ]
    _header(ws, r, 1, "Stat", bg=DGRAY, fg="000000", size=10)
    _header(ws, r, 2, "Value", bg=DGRAY, fg="000000", size=10)
    r += 1
    for label, val in stat_rows:
        _cell(ws, r, 1, label, bold=True)
        _cell(ws, r, 2, val, align="center")
        r += 1
    r += 1

    # team grades
    _subtitle(ws, r, "TEAM GRADES", 8)
    r += 1
    grade_rows = [
        ("Offense",  d.get("offense_grade")),
        ("Rotation", d.get("rotation_grade")),
        ("Bullpen",  d.get("bullpen_grade")),
        ("Team ERA", f"{d.get('team_era')}  (rank {d.get('era_rank')}/30)"),
    ]
    for label, val in grade_rows:
        _cell(ws, r, 1, label, bold=True)
        color = _grade_color(str(val)) if val else LGRAY
        _cell(ws, r, 2, val, align="center", color=color)
        r += 1
    r += 1

    # outlook
    _subtitle(ws, r, "SEASON OUTLOOK", 8)
    r += 1
    outlook_rows = [
        ("Most likely",   ou.get("most_likely")),
        ("Floor",         f"{ou.get('floor')} wins"),
        ("Ceiling",       f"{ou.get('ceiling')} wins"),
        ("Division",      ou.get("division")),
        ("Playoff path",  ou.get("playoff_path")),
        ("October",       ou.get("october_outlook")),
    ]
    for label, val in outlook_rows:
        _cell(ws, r, 1, label, bold=True)
        ws.merge_cells(start_row=r, start_column=2,
                       end_row=r, end_column=8)
        c = ws.cell(row=r, column=2, value=val)
        c.font      = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", wrap_text=True)
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    # key factors
    _subtitle(ws, r, "KEY FACTORS", 8)
    r += 1
    for f in ou.get("key_factors", []):
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=8)
        c = ws.cell(row=r, column=1, value=f"✓  {f}")
        c.font      = Font(name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=GREEN)
        c.alignment = Alignment(indent=1)
        r += 1
    r += 1

    # risks
    _subtitle(ws, r, "RISKS", 8, bg="8B0000")
    r += 1
    for risk in ou.get("risks", []):
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=8)
        c = ws.cell(row=r, column=1, value=f"⚠  {risk}")
        c.font      = Font(name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=RED)
        c.alignment = Alignment(indent=1)
        r += 1


# ── sheet 2: player grades ────────────────────────────────────────────────────
def _sheet_grades(wb: Workbook, grades: dict):
    ws = wb.create_sheet("Player Grades")
    ws.sheet_view.showGridLines = False

    r = 1
    _title(ws, r, "PLAYER GRADES", 10)
    r += 1

    # batters
    _subtitle(ws, r, "BATTERS", 10)
    r += 1
    bat_headers = ["Name","Grade","Role","PA","OPS","xwOBA",
                   "Luck","HR","WAR","Action"]
    _set_col_widths(ws, [26,22,10,6,7,7,7,5,6,45])
    for i, h in enumerate(bat_headers, 1):
        _header(ws, r, i, h, size=10)
    r += 1

    for g in grades.get("batters", []):
        color = _grade_color(g.get("grade",""))
        _cell(ws, r, 1, g["name"], bold=True, color=color)
        _cell(ws, r, 2, g["grade"], color=color)
        _cell(ws, r, 3, g.get("role",""), color=color, align="center")
        _cell(ws, r, 4, g.get("PA"), color=color, align="center")
        _cell(ws, r, 5, g.get("OPS"), color=color, align="center",
              fmt="0.000")
        _cell(ws, r, 6, g.get("xwOBA"), color=color, align="center",
              fmt="0.000")
        luck = g.get("luck")
        luck_color = GREEN if luck and luck < -0.020 else \
                     RED if luck and luck > 0.020 else color
        _cell(ws, r, 7, luck, color=luck_color, align="center", fmt="+0.000;-0.000;0.000")
        _cell(ws, r, 8, g.get("HR"), color=color, align="center")
        _cell(ws, r, 9, g.get("WAR"), color=color, align="center",
              fmt="0.0")
        _cell(ws, r, 10, g.get("action",""), color=color)
        r += 1

    r += 1
    # pitchers
    _subtitle(ws, r, "PITCHERS", 10)
    r += 1
    pit_headers = ["Name","Role","Grade","IP","ERA","WHIP",
                   "K/9","xwOBA vs","Luck","WAR","Action"]
    pit_widths = [26,5,20,7,7,7,6,8,7,6,45]
    for i, (h, w) in enumerate(zip(pit_headers, pit_widths), 1):
        _header(ws, r, i, h, size=10)
        ws.column_dimensions[get_column_letter(i)].width = w
    r += 1

    for g in grades.get("pitchers", []):
        color = _grade_color(g.get("grade",""))
        _cell(ws, r, 1,  g["name"], bold=True, color=color)
        _cell(ws, r, 2,  g.get("role",""), color=color, align="center")
        _cell(ws, r, 3,  g.get("grade",""), color=color)
        _cell(ws, r, 4,  g.get("IP"), color=color, align="center", fmt="0.0")
        _cell(ws, r, 5,  g.get("ERA"), color=color, align="center", fmt="0.00")
        _cell(ws, r, 6,  g.get("WHIP"), color=color, align="center", fmt="0.000")
        _cell(ws, r, 7,  g.get("K/9"), color=color, align="center", fmt="0.0")
        _cell(ws, r, 8,  g.get("xwOBA_against"), color=color,
              align="center", fmt="0.000")
        luck = g.get("luck")
        luck_color = GREEN if luck and luck > 0.020 else \
                     RED if luck and luck < -0.020 else color
        _cell(ws, r, 9,  luck, color=luck_color, align="center",
              fmt="+0.000;-0.000;0.000")
        _cell(ws, r, 10, g.get("WAR"), color=color, align="center", fmt="0.0")
        _cell(ws, r, 11, g.get("action",""), color=color)
        r += 1


# ── sheet 3: statcast ─────────────────────────────────────────────────────────
def _sheet_statcast(wb: Workbook, data: dict):
    ws = wb.create_sheet("Statcast")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [26,6,7,7,7,8,7,8,6])

    r = 1
    _title(ws, r, "STATCAST ANALYSIS", 9)
    r += 1

    # batters
    _subtitle(ws, r, "BATTER STATCAST", 9)
    r += 1
    headers = ["Name","PA","wOBA","xwOBA","Luck","Barrel%","HardHit%","EV50","K%"]
    for i, h in enumerate(headers, 1):
        _header(ws, r, i, h, size=10)
    r += 1

    sea_bat = data.get("statcast",{}).get("bat_luck", pd.DataFrame())
    if not sea_bat.empty:
        sea_bat_sorted = sea_bat.sort_values("xwOBA", ascending=False) \
                         if "xwOBA" in sea_bat.columns else sea_bat
        for _, row in sea_bat_sorted.iterrows():
            luck    = row.get("luck_gap")
            verdict = str(row.get("verdict",""))
            color   = GREEN if "UNLUCKY" in verdict else \
                      RED   if "LUCKY"   in verdict else LGRAY
            _cell(ws, r, 1, row.get("Name",""), bold=True)
            _cell(ws, r, 2, row.get("PA"), align="center")
            _cell(ws, r, 3, row.get("wOBA"), align="center", fmt="0.000")
            _cell(ws, r, 4, row.get("xwOBA"), align="center",
                  fmt="0.000", color=color)
            _cell(ws, r, 5, luck, align="center",
                  fmt="+0.000;-0.000;0.000", color=color)
            _cell(ws, r, 6, row.get("Barrel%"), align="center", fmt="0.0")
            _cell(ws, r, 7, row.get("HardHit%"), align="center", fmt="0.0")
            _cell(ws, r, 8, row.get("EV50"), align="center", fmt="0.0")
            _cell(ws, r, 9, row.get("K%"), align="center", fmt="0.0")
            r += 1

    r += 1
    # pitchers
    _subtitle(ws, r, "PITCHER STATCAST", 9)
    r += 1
    pit_headers = ["Name","BF","wOBA vs","xwOBA vs","Luck","K%","Whiff%","HardHit% vs","Barrel% vs"]
    for i, h in enumerate(pit_headers, 1):
        _header(ws, r, i, h, size=10)
    r += 1

    sea_pit = data.get("statcast",{}).get("pit_luck", pd.DataFrame())
    if not sea_pit.empty:
        xwa = "xwOBA_against"
        sea_pit_sorted = sea_pit.sort_values(xwa) \
                         if xwa in sea_pit.columns else sea_pit
        for _, row in sea_pit_sorted.iterrows():
            luck    = row.get("luck_gap")
            verdict = str(row.get("verdict",""))
            color   = GREEN if "UNLUCKY" in verdict else \
                      RED   if "LUCKY"   in verdict else LGRAY
            _cell(ws, r, 1, row.get("Name",""), bold=True)
            _cell(ws, r, 2, row.get("PA",row.get("BF")), align="center")
            _cell(ws, r, 3, row.get("wOBA_against"), align="center", fmt="0.000")
            _cell(ws, r, 4, row.get(xwa), align="center",
                  fmt="0.000", color=color)
            _cell(ws, r, 5, luck, align="center",
                  fmt="+0.000;-0.000;0.000", color=color)
            _cell(ws, r, 6, row.get("K%"), align="center", fmt="0.0")
            _cell(ws, r, 7, row.get("Whiff%"), align="center", fmt="0.0")
            _cell(ws, r, 8, row.get("HardHit%_against",
                                    row.get("HardHit%")), align="center", fmt="0.0")
            _cell(ws, r, 9, row.get("Barrel%_against",
                                    row.get("Barrel%")), align="center", fmt="0.0")
            r += 1


# ── sheet 4: stats to improve ─────────────────────────────────────────────────
def _sheet_stats(wb: Workbook, stats: list):
    ws = wb.create_sheet("Stats to Improve")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [12,12,10,8,8,12,50])

    r = 1
    _title(ws, r, "STATS TO IMPROVE", 7)
    r += 1

    headers = ["Priority","Category","Stat","Value","Rank","Internal Fix?","Recommended Fix"]
    for i, h in enumerate(headers, 1):
        _header(ws, r, i, h, size=10)
    r += 1

    for s in stats:
        color = RED if s["priority"] == "HIGH" else YELLOW
        _cell(ws, r, 1, s["priority"], bold=True, color=color, align="center")
        _cell(ws, r, 2, s["category"], color=color)
        _cell(ws, r, 3, s["stat"], bold=True, color=color)
        _cell(ws, r, 4, s["value"], color=color, align="center")
        _cell(ws, r, 5, f"{s['rank']}/{s['total']}", color=color, align="center")
        internal = "✓ Yes" if s["internal"] else "→ External"
        _cell(ws, r, 6, internal, color=GREEN if s["internal"] else ORANGE,
              align="center")
        fix_cell = ws.cell(row=r, column=7, value=s["fix"])
        fix_cell.font      = Font(name="Arial", size=10)
        fix_cell.fill      = PatternFill("solid", start_color=color)
        fix_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 30
        r += 1


# ── sheet 5: roster moves ─────────────────────────────────────────────────────
def _sheet_moves(wb: Workbook, im: list, dl: dict):
    ws = wb.create_sheet("Roster Moves")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [14,14,28,50])

    r = 1
    _title(ws, r, "ROSTER MOVES", 4)
    r += 1

    # immediate
    _subtitle(ws, r, "IMMEDIATE MOVES", 4)
    r += 1
    for i, h in enumerate(["Urgency","Type","Player","Reason"], 1):
        _header(ws, r, i, h, size=10)
    r += 1

    URGENCY_COLOR = {
        "IMMEDIATE":  RED,
        "SOON":       ORANGE,
        "WHEN READY": YELLOW,
    }
    for m in im:
        color = URGENCY_COLOR.get(m["urgency"], LGRAY)
        _cell(ws, r, 1, m["urgency"], bold=True, color=color, align="center")
        _cell(ws, r, 2, m["type"], color=color, align="center")
        _cell(ws, r, 3, m["player"], bold=True, color=color)
        _cell(ws, r, 4, m["reason"], color=color)
        r += 1

    r += 1
    # deadline
    _subtitle(ws, r, "DEADLINE STRATEGY", 4)
    r += 1
    ws.merge_cells(start_row=r, start_column=1,
                   end_row=r, end_column=4)
    c = ws.cell(row=r, column=1, value=dl.get("strategy",""))
    c.font      = Font(name="Arial", size=10, italic=True)
    c.fill      = PatternFill("solid", start_color=LGRAY)
    c.alignment = Alignment(wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 40
    r += 2

    # targets
    _subtitle(ws, r, "DEADLINE TARGETS", 4, bg=TEAL)
    r += 1
    for t in dl.get("targets", []):
        items = [
            ("Position",  t.get("position","")),
            ("Profile",   t.get("profile","")),
            ("Cost",      t.get("cost","")),
            ("Urgency",   t.get("urgency","")),
            ("Why",       t.get("why","")),
        ]
        for label, val in items:
            _cell(ws, r, 1, label, bold=True, color=LGRAY)
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r, end_column=4)
            c = ws.cell(row=r, column=2, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", start_color=LGRAY)
            c.alignment = Alignment(wrap_text=True)
            r += 1
        r += 1

    # do not trade
    _subtitle(ws, r, "DO NOT TRADE", 4, bg="8B0000")
    r += 1
    for p in dl.get("do_not_trade", []):
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=f"✗  {p}")
        c.font      = Font(name="Arial", size=10, bold=True)
        c.fill      = PatternFill("solid", start_color=RED)
        c.alignment = Alignment(indent=1)
        r += 1

    r += 1
    # sell candidates
    if dl.get("sell"):
        _subtitle(ws, r, "SELL CANDIDATES", 4, bg=GOLD)
        r += 1
        for s in dl["sell"]:
            _cell(ws, r, 1, s["player"], bold=True, color=YELLOW)
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r, end_column=4)
            c = ws.cell(row=r, column=2,
                        value=f"{s['reason']}  |  Note: {s.get('note','')}")
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", start_color=YELLOW)
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 30
            r += 1


# ── sheet 6: schedule ─────────────────────────────────────────────────────────
def _sheet_schedule(wb: Workbook, data: dict):
    ws = wb.create_sheet("Schedule")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [6,12,8,12,10,10,10,10])

    r = 1
    _title(ws, r, "SCHEDULE", 8)
    r += 1

    # next 7
    _subtitle(ws, r, "NEXT 7 GAMES", 8)
    r += 1
    for i, h in enumerate(["#","Date","H/A","Opponent",
                            "Result","W","L","Notes"], 1):
        _header(ws, r, i, h, size=10)
    r += 1

    next7 = data.get("schedule",{}).get("next7", pd.DataFrame())
    if next7 is not None and not next7.empty:
        for _, row in next7.iterrows():
            checked = row.get("checked", False)
            color   = GREEN if checked else LGRAY
            _cell(ws, r, 1, "☑" if checked else "☐",
                  align="center", color=color)
            _cell(ws, r, 2, str(row.get("date",""))[:10], color=color)
            _cell(ws, r, 3, row.get("home_away",""), align="center", color=color)
            _cell(ws, r, 4, row.get("Opp",""), color=color)
            _cell(ws, r, 5, row.get("W-L",""), align="center", color=color)
            _cell(ws, r, 6, row.get("W",""), align="center", color=color)
            _cell(ws, r, 7, row.get("L",""), align="center", color=color)
            _cell(ws, r, 8, "", color=color)
            r += 1

    r += 1
    # remaining
    _subtitle(ws, r, "REMAINING SCHEDULE", 8)
    r += 1
    for i, h in enumerate(["#","Date","H/A","Opponent",
                            "Result","W","L","Notes"], 1):
        _header(ws, r, i, h, size=10)
    r += 1

    remaining = data.get("schedule",{}).get("remaining", pd.DataFrame())
    if remaining is not None and not remaining.empty:
        bad_teams  = ["LAA","HOU","COL","MIA","DET","KC",
                      "BAL","BOS","TOR","WSN"]
        good_teams = ["TBR","NYY","LAD","MIL","CHC","ATL"]
        for _, row in remaining.iterrows():
            opp   = str(row.get("Opp",""))
            color = GREEN  if any(t in opp for t in bad_teams)  else \
                    RED    if any(t in opp for t in good_teams)  else \
                    LGRAY
            _cell(ws, r, 1, row.get("Gm#",""), align="center", color=color)
            _cell(ws, r, 2, str(row.get("date",""))[:10], color=color)
            _cell(ws, r, 3, row.get("home_away",""), align="center", color=color)
            _cell(ws, r, 4, opp, color=color)
            _cell(ws, r, 5, row.get("W-L",""), align="center", color=color)
            _cell(ws, r, 6, row.get("W",""), align="center", color=color)
            _cell(ws, r, 7, row.get("L",""), align="center", color=color)
            _cell(ws, r, 8, "", color=color)
            r += 1



# ── sheet 7: player targets ───────────────────────────────────────────────────
def _sheet_targets(wb: Workbook, targets: dict):
    ws = wb.create_sheet("Trade Targets")
    ws.sheet_view.showGridLines = False

    r = 1
    _title(ws, r, "MLB TRADE & WAIVER TARGETS", 9)
    r += 1

    # batter targets
    _subtitle(ws, r, "1B/DH TARGETS — xwOBA .330+, Barrel% 8%+, non-contender", 9)
    r += 1
    bat_h = ["Name","Team","PA","xwOBA","Barrel%","HardHit%","HR","OPS","Fit"]
    bat_w = [25,7,6,8,9,10,5,7,30]
    _set_col_widths(ws, bat_w)
    for i,(h,w) in enumerate(zip(bat_h,bat_w),1):
        _header(ws, r, i, h, size=10)
        ws.column_dimensions[get_column_letter(i)].width = w
    r += 1

    for p in targets.get("batter_targets",[])[:10]:
        xw = pd.to_numeric(p.get("xwOBA"), errors="coerce")
        color = GREEN if xw and xw >= 0.360 else                 "E8F5E9" if xw and xw >= 0.340 else LGRAY
        _cell(ws, r, 1, p.get("name",""), bold=True, color=color)
        _cell(ws, r, 2, p.get("team",""), color=color, align="center")
        _cell(ws, r, 3, p.get("PA"), color=color, align="center")
        _cell(ws, r, 4, p.get("xwOBA"), color=color, align="center", fmt="0.000")
        _cell(ws, r, 5, p.get("Barrel%"), color=color, align="center", fmt="0.0")
        _cell(ws, r, 6, p.get("HardHit%"), color=color, align="center", fmt="0.0")
        _cell(ws, r, 7, p.get("HR"), color=color, align="center")
        _cell(ws, r, 8, p.get("OPS"), color=color, align="center", fmt="0.000")
        _cell(ws, r, 9, p.get("fit",""), color=color)
        r += 1

    r += 1
    # pitcher targets
    _subtitle(ws, r, "BULLPEN TARGETS — xwOBA against ≤.310, 20+ IP, non-contender", 9)
    r += 1
    pit_h = ["Name","Team","G","IP","ERA","WHIP","K%","xwOBA vs","Fit"]
    pit_w = [25,7,5,6,7,7,6,9,30]
    for i,(h,w) in enumerate(zip(pit_h,pit_w),1):
        _header(ws, r, i, h, size=10)
        ws.column_dimensions[get_column_letter(i)].width = w
    r += 1

    for p in targets.get("pitcher_targets",[])[:10]:
        xw = pd.to_numeric(p.get("xwOBA_against"), errors="coerce")
        color = GREEN if xw and xw <= 0.270 else                 "E8F5E9" if xw and xw <= 0.290 else LGRAY
        _cell(ws, r, 1, p.get("name",""), bold=True, color=color)
        _cell(ws, r, 2, p.get("team",""), color=color, align="center")
        _cell(ws, r, 3, p.get("G"), color=color, align="center")
        _cell(ws, r, 4, p.get("IP"), color=color, align="center", fmt="0.0")
        _cell(ws, r, 5, p.get("ERA"), color=color, align="center", fmt="0.00")
        _cell(ws, r, 6, p.get("WHIP"), color=color, align="center", fmt="0.000")
        _cell(ws, r, 7, p.get("K%"), color=color, align="center", fmt="0.0")
        _cell(ws, r, 8, p.get("xwOBA_against"), color=color, align="center", fmt="0.000")
        _cell(ws, r, 9, p.get("fit",""), color=color)
        r += 1

# ── main generate function ────────────────────────────────────────────────────
def generate_report(data: dict, analysis: dict,
                    grades: dict, recs: dict,
                    output_path: str = None) -> str:
    """
    Generates the Excel report.
    Returns path to the saved file.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = output_path or OUTPUT_FILE

    print(f"\n[report] Generating Excel report...")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    d  = recs["diagnosis"]
    ou = recs["outlook"]
    si = recs["stats_to_improve"]
    im = recs["immediate_moves"]
    dl = recs["deadline"]

    _sheet_summary(wb, d, ou)
    _sheet_grades(wb, grades)
    _sheet_statcast(wb, data)
    _sheet_stats(wb, si)
    _sheet_moves(wb, im, dl)
    _sheet_targets(wb, recs.get("player_targets", {}))
    _sheet_schedule(wb, data)

    wb.save(path)
    print(f"[report] Saved: {path}")
    return path


# ── test ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    from data_builder import build_all
    from team_analyzer import analyze_team
    from player_grades import grade_players
    from recommender   import generate_recommendations

    data     = build_all(2026)
    analysis = analyze_team(data)
    grades   = grade_players(data)
    recs     = generate_recommendations(data, analysis, grades)
    path     = generate_report(data, analysis, grades, recs)
    print(f"\nReport saved to: {path}")